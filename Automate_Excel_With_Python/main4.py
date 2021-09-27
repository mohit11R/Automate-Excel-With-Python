from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter

# LOADING SHEET
wb = load_workbook('testing.xlsx')
ws = wb.active

# Merging Cells

ws.merge_cells('A1:D1')


# UnMerging Cells

ws.unmerge_cells('A1:D1')

wb.save('testing.xlsx')