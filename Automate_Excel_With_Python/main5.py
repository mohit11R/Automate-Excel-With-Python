from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter

# LOADING SHEET
wb = load_workbook('testing.xlsx')
ws = wb.active

# Insert Row
ws.insert_rows(7)
ws.insert_rows(7)

# Delete Row

ws.delete_rows(7)

# Insest Column
ws.insert_cols(2)

# Delete Column
ws.delete_cols(2)

# Copying values and Moving them 
ws.move_range("C1:D11",rows=2,cols=2)

wb.save('testing.xlsx')