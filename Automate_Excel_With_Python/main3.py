from openpyxl import Workbook , load_workbook
from openpyxl.utils import get_column_letter

# LOADING SHEET
wb = load_workbook('testing.xlsx')
ws = wb.active

# Accessing Multiple values from sheet
for row in range(1,11):
    for col in range(1,5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)