from openpyxl import Workbook, load_workbook


wb = load_workbook('Grades.xlsx')

ws = wb.active

# Accessing the value from sheet
print(ws['A1'].value)

# Editing a cell value 

# ws['A1'] = 'CPP'
# wb.save('Grades.xlsx')


# list of sheet in wb 
print(wb.sheetnames)


# Creating new Sheet

wb.create_sheet('test')
print(wb.sheetnames)